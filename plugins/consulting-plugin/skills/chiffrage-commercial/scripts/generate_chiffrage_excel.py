#!/usr/bin/env python3
"""
Script de génération et remplissage du template Excel de chiffrage
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
from typing import Dict, List, Any


class ChiffrageExcelGenerator:
    """Générateur de fichier Excel de chiffrage"""
    
    def __init__(self, chiffrage_data: Dict[str, Any], profiles_data: Dict[str, Any]):
        """
        Initialise le générateur
        
        Args:
            chiffrage_data: Données du chiffrage (phases, activités, charges)
            profiles_data: Données des profils (TJM, etc.)
        """
        self.wb = openpyxl.Workbook()
        self.chiffrage_data = chiffrage_data
        self.profiles_data = profiles_data
        
        # Styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.subheader_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        self.total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        self.total_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def apply_header_style(self, cell):
        """Applique le style header à une cellule"""
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border
    
    def apply_subheader_style(self, cell):
        """Applique le style subheader à une cellule"""
        cell.fill = self.subheader_fill
        cell.font = self.subheader_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.border
    
    def apply_total_style(self, cell):
        """Applique le style total à une cellule"""
        cell.fill = self.total_fill
        cell.font = self.total_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = self.border
    
    def apply_border(self, cell):
        """Applique une bordure à une cellule"""
        cell.border = self.border
    
    def create_hypotheses_sheet(self):
        """Crée l'onglet Hypothèses"""
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        
        ws = self.wb.create_sheet("Hypothèses", 0)
        
        # En-tête
        ws['A1'] = "HYPOTHÈSES DE CHIFFRAGE"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # Informations générales
        row = 3
        ws[f'A{row}'] = "Mission"
        ws[f'B{row}'] = self.chiffrage_data.get('mission_name', '')
        self.apply_subheader_style(ws[f'A{row}'])
        
        row += 1
        ws[f'A{row}'] = "Type"
        ws[f'B{row}'] = self.chiffrage_data.get('mission_type', '')
        
        row += 1
        ws[f'A{row}'] = "Client"
        ws[f'B{row}'] = self.chiffrage_data.get('client', '')
        
        row += 1
        ws[f'A{row}'] = "Date"
        ws[f'B{row}'] = datetime.now().strftime("%d/%m/%Y")
        
        # Hypothèses structurelles
        row += 2
        ws[f'A{row}'] = "HYPOTHÈSES STRUCTURELLES"
        self.apply_subheader_style(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        for hypothesis in self.chiffrage_data.get('hypotheses', []):
            ws[f'A{row}'] = f"• {hypothesis}"
            row += 1
        
        # Taux journaliers
        row += 2
        ws[f'A{row}'] = "TAUX JOURNALIERS MOYENS (TJM)"
        self.apply_subheader_style(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Profil"
        ws[f'B{row}'] = "TJM (€)"
        ws[f'C{row}'] = "Description"
        self.apply_header_style(ws[f'A{row}'])
        self.apply_header_style(ws[f'B{row}'])
        self.apply_header_style(ws[f'C{row}'])
        
        for profile in self.profiles_data['profiles']:
            row += 1
            ws[f'A{row}'] = profile['name']
            ws[f'B{row}'] = profile['tjm']
            ws[f'B{row}'].number_format = '#,##0 €'
            ws[f'C{row}'] = profile['description']
            self.apply_border(ws[f'A{row}'])
            self.apply_border(ws[f'B{row}'])
            self.apply_border(ws[f'C{row}'])
        
        # Ajustement des largeurs
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
    
    def create_charges_sheet(self):
        """Crée l'onglet Charges (en jours)"""
        ws = self.wb.create_sheet("Charges (jours)")
        
        # En-tête
        ws['A1'] = "CHARGES PAR ACTIVITÉ ET PAR PROFIL (en jours)"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:H1')
        
        # Construction de l'en-tête des colonnes
        profiles = self.profiles_data['profiles']
        
        row = 3
        ws[f'A{row}'] = "Phase"
        ws[f'B{row}'] = "Activité"
        self.apply_header_style(ws[f'A{row}'])
        self.apply_header_style(ws[f'B{row}'])
        
        col = 3  # Colonne C
        for profile in profiles:
            cell = ws.cell(row=row, column=col)
            cell.value = profile['name']
            self.apply_header_style(cell)
            col += 1
        
        # Total
        cell = ws.cell(row=row, column=col)
        cell.value = "Total (jours)"
        self.apply_header_style(cell)
        
        # Remplissage des données
        row += 1
        start_data_row = row
        
        for phase in self.chiffrage_data.get('phases', []):
            phase_start_row = row
            
            for activity in phase.get('activities', []):
                ws.cell(row=row, column=1).value = phase['name']
                ws.cell(row=row, column=2).value = activity['name']
                self.apply_border(ws.cell(row=row, column=1))
                self.apply_border(ws.cell(row=row, column=2))
                
                col = 3
                total_formula_parts = []
                for profile in profiles:
                    cell = ws.cell(row=row, column=col)
                    # Récupérer la charge pour ce profil
                    charge = activity.get('charges', {}).get(profile['id'], 0)
                    cell.value = charge if charge > 0 else ""
                    cell.number_format = '0.0'
                    self.apply_border(cell)
                    
                    if charge > 0:
                        total_formula_parts.append(get_column_letter(col) + str(row))
                    
                    col += 1
                
                # Total de la ligne
                total_cell = ws.cell(row=row, column=col)
                if total_formula_parts:
                    total_cell.value = f"=SUM({','.join(total_formula_parts)})"
                    total_cell.number_format = '0.0'
                self.apply_border(total_cell)
                
                row += 1
            
            # Sous-total de phase
            ws.cell(row=row, column=1).value = f"Sous-total {phase['name']}"
            ws.cell(row=row, column=2).value = ""
            self.apply_subheader_style(ws.cell(row=row, column=1))
            self.apply_subheader_style(ws.cell(row=row, column=2))
            
            col = 3
            for _ in profiles:
                cell = ws.cell(row=row, column=col)
                col_letter = get_column_letter(col)
                cell.value = f"=SUM({col_letter}{phase_start_row}:{col_letter}{row-1})"
                cell.number_format = '0.0'
                self.apply_subheader_style(cell)
                col += 1
            
            # Total sous-total
            total_cell = ws.cell(row=row, column=col)
            col_letter = get_column_letter(col)
            total_cell.value = f"=SUM({col_letter}{phase_start_row}:{col_letter}{row-1})"
            total_cell.number_format = '0.0'
            self.apply_subheader_style(total_cell)
            
            row += 1
        
        # Total général
        row += 1
        ws.cell(row=row, column=1).value = "TOTAL GÉNÉRAL"
        ws.cell(row=row, column=2).value = ""
        self.apply_total_style(ws.cell(row=row, column=1))
        self.apply_total_style(ws.cell(row=row, column=2))
        
        col = 3
        for _ in profiles:
            cell = ws.cell(row=row, column=col)
            col_letter = get_column_letter(col)
            cell.value = f"=SUM({col_letter}{start_data_row}:{col_letter}{row-1})"
            cell.number_format = '0.0'
            self.apply_total_style(cell)
            col += 1
        
        # Total total
        total_cell = ws.cell(row=row, column=col)
        col_letter = get_column_letter(col)
        total_cell.value = f"=SUM({col_letter}{start_data_row}:{col_letter}{row-1})"
        total_cell.number_format = '0.0'
        self.apply_total_style(total_cell)
        
        # Ajustement des largeurs
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        for i in range(3, col + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15
    
    def create_couts_sheet(self):
        """Crée l'onglet Coûts (en €)"""
        ws = self.wb.create_sheet("Coûts (€)")
        
        # En-tête
        ws['A1'] = "COÛTS PAR ACTIVITÉ ET PAR PROFIL (en €)"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:H1')
        
        # Construction de l'en-tête des colonnes
        profiles = self.profiles_data['profiles']
        
        row = 3
        ws[f'A{row}'] = "Phase"
        ws[f'B{row}'] = "Activité"
        self.apply_header_style(ws[f'A{row}'])
        self.apply_header_style(ws[f'B{row}'])
        
        col = 3
        for profile in profiles:
            cell = ws.cell(row=row, column=col)
            cell.value = profile['name']
            self.apply_header_style(cell)
            col += 1
        
        # Total
        cell = ws.cell(row=row, column=col)
        cell.value = "Total (€)"
        self.apply_header_style(cell)
        
        # Remplissage des données
        row += 1
        start_data_row = row
        
        for phase in self.chiffrage_data.get('phases', []):
            phase_start_row = row
            
            for activity in phase.get('activities', []):
                ws.cell(row=row, column=1).value = phase['name']
                ws.cell(row=row, column=2).value = activity['name']
                self.apply_border(ws.cell(row=row, column=1))
                self.apply_border(ws.cell(row=row, column=2))
                
                col = 3
                total_formula_parts = []
                for profile in profiles:
                    cell = ws.cell(row=row, column=col)
                    # Calculer le coût : charge * TJM
                    charge = activity.get('charges', {}).get(profile['id'], 0)
                    if charge > 0:
                        cost = charge * profile['tjm']
                        cell.value = cost
                        cell.number_format = '#,##0 €'
                        total_formula_parts.append(get_column_letter(col) + str(row))
                    else:
                        cell.value = ""
                    self.apply_border(cell)
                    
                    col += 1
                
                # Total de la ligne
                total_cell = ws.cell(row=row, column=col)
                if total_formula_parts:
                    total_cell.value = f"=SUM({','.join(total_formula_parts)})"
                    total_cell.number_format = '#,##0 €'
                self.apply_border(total_cell)
                
                row += 1
            
            # Sous-total de phase
            ws.cell(row=row, column=1).value = f"Sous-total {phase['name']}"
            ws.cell(row=row, column=2).value = ""
            self.apply_subheader_style(ws.cell(row=row, column=1))
            self.apply_subheader_style(ws.cell(row=row, column=2))
            
            col = 3
            for _ in profiles:
                cell = ws.cell(row=row, column=col)
                col_letter = get_column_letter(col)
                cell.value = f"=SUM({col_letter}{phase_start_row}:{col_letter}{row-1})"
                cell.number_format = '#,##0 €'
                self.apply_subheader_style(cell)
                col += 1
            
            # Total sous-total
            total_cell = ws.cell(row=row, column=col)
            col_letter = get_column_letter(col)
            total_cell.value = f"=SUM({col_letter}{phase_start_row}:{col_letter}{row-1})"
            total_cell.number_format = '#,##0 €'
            self.apply_subheader_style(total_cell)
            
            row += 1
        
        # Total général
        row += 1
        ws.cell(row=row, column=1).value = "TOTAL GÉNÉRAL"
        ws.cell(row=row, column=2).value = ""
        self.apply_total_style(ws.cell(row=row, column=1))
        self.apply_total_style(ws.cell(row=row, column=2))
        
        col = 3
        for _ in profiles:
            cell = ws.cell(row=row, column=col)
            col_letter = get_column_letter(col)
            cell.value = f"=SUM({col_letter}{start_data_row}:{col_letter}{row-1})"
            cell.number_format = '#,##0 €'
            self.apply_total_style(cell)
            col += 1
        
        # Total total
        total_cell = ws.cell(row=row, column=col)
        col_letter = get_column_letter(col)
        total_cell.value = f"=SUM({col_letter}{start_data_row}:{col_letter}{row-1})"
        total_cell.number_format = '#,##0 €'
        self.apply_total_style(total_cell)
        
        # Ajustement des largeurs
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        for i in range(3, col + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15
    
    def create_synthese_sheet(self):
        """Crée l'onglet Synthèse"""
        ws = self.wb.create_sheet("Synthèse")
        
        # En-tête
        ws['A1'] = "SYNTHÈSE DU CHIFFRAGE"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Récapitulatif par phase
        ws[f'A{row}'] = "RÉCAPITULATIF PAR PHASE"
        self.apply_subheader_style(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = "Phase"
        ws[f'B{row}'] = "Charges (jours)"
        ws[f'C{row}'] = "Coûts (€)"
        ws[f'D{row}'] = "% du total"
        self.apply_header_style(ws[f'A{row}'])
        self.apply_header_style(ws[f'B{row}'])
        self.apply_header_style(ws[f'C{row}'])
        self.apply_header_style(ws[f'D{row}'])
        
        total_charges = 0
        total_couts = 0
        
        for phase in self.chiffrage_data.get('phases', []):
            row += 1
            phase_charges = sum(
                sum(activity.get('charges', {}).values())
                for activity in phase.get('activities', [])
            )
            phase_couts = sum(
                sum(
                    charge * next((p['tjm'] for p in self.profiles_data['profiles'] if p['id'] == prof_id), 0)
                    for prof_id, charge in activity.get('charges', {}).items()
                )
                for activity in phase.get('activities', [])
            )
            
            total_charges += phase_charges
            total_couts += phase_couts
            
            ws[f'A{row}'] = phase['name']
            ws[f'B{row}'] = phase_charges
            ws[f'B{row}'].number_format = '0.0'
            ws[f'C{row}'] = phase_couts
            ws[f'C{row}'].number_format = '#,##0 €'
            ws[f'D{row}'] = phase_couts / total_couts if total_couts > 0 else 0
            ws[f'D{row}'].number_format = '0.0%'
            
            self.apply_border(ws[f'A{row}'])
            self.apply_border(ws[f'B{row}'])
            self.apply_border(ws[f'C{row}'])
            self.apply_border(ws[f'D{row}'])
        
        # Total
        row += 1
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = total_charges
        ws[f'B{row}'].number_format = '0.0'
        ws[f'C{row}'] = total_couts
        ws[f'C{row}'].number_format = '#,##0 €'
        ws[f'D{row}'] = "100%"
        self.apply_total_style(ws[f'A{row}'])
        self.apply_total_style(ws[f'B{row}'])
        self.apply_total_style(ws[f'C{row}'])
        self.apply_total_style(ws[f'D{row}'])
        
        # Récapitulatif par profil
        row += 3
        ws[f'A{row}'] = "RÉCAPITULATIF PAR PROFIL"
        self.apply_subheader_style(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = "Profil"
        ws[f'B{row}'] = "Charges (jours)"
        ws[f'C{row}'] = "Coûts (€)"
        ws[f'D{row}'] = "% du total"
        self.apply_header_style(ws[f'A{row}'])
        self.apply_header_style(ws[f'B{row}'])
        self.apply_header_style(ws[f'C{row}'])
        self.apply_header_style(ws[f'D{row}'])
        
        for profile in self.profiles_data['profiles']:
            row += 1
            profile_charges = sum(
                sum(
                    activity.get('charges', {}).get(profile['id'], 0)
                    for activity in phase.get('activities', [])
                )
                for phase in self.chiffrage_data.get('phases', [])
            )
            profile_couts = profile_charges * profile['tjm']
            
            ws[f'A{row}'] = profile['name']
            ws[f'B{row}'] = profile_charges
            ws[f'B{row}'].number_format = '0.0'
            ws[f'C{row}'] = profile_couts
            ws[f'C{row}'].number_format = '#,##0 €'
            ws[f'D{row}'] = profile_couts / total_couts if total_couts > 0 else 0
            ws[f'D{row}'].number_format = '0.0%'
            
            self.apply_border(ws[f'A{row}'])
            self.apply_border(ws[f'B{row}'])
            self.apply_border(ws[f'C{row}'])
            self.apply_border(ws[f'D{row}'])
        
        # Total
        row += 1
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = total_charges
        ws[f'B{row}'].number_format = '0.0'
        ws[f'C{row}'] = total_couts
        ws[f'C{row}'].number_format = '#,##0 €'
        ws[f'D{row}'] = "100%"
        self.apply_total_style(ws[f'A{row}'])
        self.apply_total_style(ws[f'B{row}'])
        self.apply_total_style(ws[f'C{row}'])
        self.apply_total_style(ws[f'D{row}'])
        
        # Indicateurs clés
        row += 3
        ws[f'A{row}'] = "INDICATEURS CLÉS"
        self.apply_subheader_style(ws[f'A{row}'])
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = "Charge totale"
        ws[f'B{row}'] = total_charges
        ws[f'B{row}'].number_format = '0.0'
        ws[f'C{row}'] = "jours"
        
        row += 1
        ws[f'A{row}'] = "Coût total"
        ws[f'B{row}'] = total_couts
        ws[f'B{row}'].number_format = '#,##0 €'
        ws[f'C{row}'] = "HT"
        
        row += 1
        ws[f'A{row}'] = "TJM moyen"
        ws[f'B{row}'] = total_couts / total_charges if total_charges > 0 else 0
        ws[f'B{row}'].number_format = '#,##0 €'
        
        # Ajustement des largeurs
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
    
    def generate(self, output_path: str):
        """
        Génère le fichier Excel complet
        
        Args:
            output_path: Chemin du fichier de sortie
        """
        self.create_hypotheses_sheet()
        self.create_charges_sheet()
        self.create_couts_sheet()
        self.create_synthese_sheet()
        
        self.wb.save(output_path)
        print(f"✅ Fichier Excel généré : {output_path}")


def generate_chiffrage_excel(chiffrage_data: Dict[str, Any], 
                             profiles_data: Dict[str, Any],
                             output_path: str):
    """
    Fonction principale de génération du fichier Excel
    
    Args:
        chiffrage_data: Données du chiffrage
        profiles_data: Données des profils
        output_path: Chemin du fichier de sortie
    """
    generator = ChiffrageExcelGenerator(chiffrage_data, profiles_data)
    generator.generate(output_path)


if __name__ == "__main__":
    # Exemple d'utilisation
    import sys
    
    if len(sys.argv) < 4:
        print("Usage: python generate_chiffrage_excel.py <chiffrage.json> <profiles.json> <output.xlsx>")
        sys.exit(1)
    
    chiffrage_file = sys.argv[1]
    profiles_file = sys.argv[2]
    output_file = sys.argv[3]
    
    with open(chiffrage_file, 'r', encoding='utf-8') as f:
        chiffrage_data = json.load(f)
    
    with open(profiles_file, 'r', encoding='utf-8') as f:
        profiles_data = json.load(f)
    
    generate_chiffrage_excel(chiffrage_data, profiles_data, output_file)
