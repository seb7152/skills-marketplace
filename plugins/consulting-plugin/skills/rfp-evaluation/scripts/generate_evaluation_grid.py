"""
Generate RFP Evaluation Grid — Excel Output
============================================

Generates a comparative evaluation Excel file from RFP evaluation data.
Data is expected as a JSON file exported from rfp-analyzer MCP.

Usage:
    python generate_evaluation_grid.py data.json output.xlsx [--format A|B]

Arguments:
    data.json     JSON file with evaluation data (categories, requirements, responses)
    output.xlsx   Output Excel file path
    --format A    One sheet per supplier + summary sheet (default)
    --format B    Single sheet, all suppliers as columns

Input JSON format:
    {
      "project": "Project Name",
      "suppliers": ["Supplier A", "Supplier B", "Supplier C"],
      "categories": [
        {"id": "CAT1", "code": "1", "title": "Functional", "level": 1, "parent_id": null}
      ],
      "requirements": [
        {
          "code": "EX-001",
          "title": "User Authentication",
          "description": "...",
          "weight": 0.8,
          "category_name": "1",
          "is_mandatory": true
        }
      ],
      "responses": {
        "Supplier A": [
          {
            "requirement_id_external": "EX-001",
            "response_text": "...",
            "ai_score": 4.0,
            "ai_comment": "...",
            "status": "pass",
            "question": ""
          }
        ]
      }
    }

Created: 2025
"""

import json
import sys
import argparse
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ─── Color palette ────────────────────────────────────────────────────────────

COLORS = {
    "header_dark":   "1E3A5F",   # Dark blue — main headers
    "header_mid":    "2E6DA4",   # Mid blue — category headers
    "header_light":  "D0E4F7",   # Light blue — column headers
    "mandatory_bg":  "FFE0E0",   # Light red — mandatory requirements
    "score_0_1":     "FF6B6B",   # Red — fail
    "score_2":       "FFA94D",   # Orange — partial
    "score_3":       "FFD43B",   # Yellow — satisfactory
    "score_4":       "69DB7C",   # Green — good
    "score_5":       "2F9E44",   # Dark green — excellent
    "pending":       "D9D9D9",   # Grey — pending
    "white":         "FFFFFF",
    "row_alt":       "F8FBFF",   # Very light blue — alternating rows
    "summary_total": "E8F0FE",   # Pale blue — totals
    "alert_red":     "FF0000",   # Red text for alerts
}


def score_fill(score):
    """Return fill color based on score value."""
    if score is None:
        return PatternFill("solid", fgColor=COLORS["pending"])
    if score <= 1:
        return PatternFill("solid", fgColor=COLORS["score_0_1"])
    if score <= 2:
        return PatternFill("solid", fgColor=COLORS["score_2"])
    if score <= 2.9:
        return PatternFill("solid", fgColor=COLORS["score_3"])
    if score <= 4:
        return PatternFill("solid", fgColor=COLORS["score_4"])
    return PatternFill("solid", fgColor=COLORS["score_5"])


def status_label(status, score, is_mandatory):
    """Return human-readable status label."""
    if status == "pending" or score is None:
        return "En attente"
    if status == "fail" or (is_mandatory and score is not None and score < 3):
        return "❌ Non conforme"
    if status == "partial":
        return "⚠️ Partiel"
    if status == "pass":
        return "✅ Conforme"
    return status or ""


def weighted_score(requirements, responses_by_req):
    """
    Calculate weighted average score for a set of requirements and their responses.
    Ignores pending (null) scores.
    """
    total_weight = 0
    weighted_sum = 0
    evaluated = 0

    for req in requirements:
        resp = responses_by_req.get(req["code"])
        if resp and resp.get("ai_score") is not None:
            w = req.get("weight", 0.5)
            weighted_sum += resp["ai_score"] * w
            total_weight += w
            evaluated += 1

    if total_weight == 0:
        return None, 0, evaluated
    return round(weighted_sum / total_weight, 2), total_weight, evaluated


def build_category_tree(categories):
    """Build hierarchical category structure."""
    cat_by_id = {c["id"]: c for c in categories}
    children = defaultdict(list)

    for cat in categories:
        parent = cat.get("parent_id")
        if parent:
            children[parent].append(cat)

    roots = [c for c in categories if not c.get("parent_id")]
    roots.sort(key=lambda c: (c.get("order", 999), c.get("code", "")))

    def sort_children(cat_list):
        return sorted(cat_list, key=lambda c: (c.get("order", 999), c.get("code", "")))

    def flatten(cats, depth=0):
        result = []
        for cat in sort_children(cats):
            result.append((cat, depth))
            result.extend(flatten(children[cat["id"]], depth + 1))
        return result

    return flatten(roots), children


def get_reqs_for_category(cat_id, requirements, categories, children_map):
    """Get all requirements belonging to a category (direct children only for leaf cats)."""
    cat_by_id = {c["id"]: c for c in categories}
    cat = cat_by_id.get(cat_id)
    if not cat:
        return []

    cat_codes = {cat["code"], cat["title"], cat["id"]}

    return [
        r for r in requirements
        if r.get("category_name") in cat_codes
    ]


# ─── Format A: One sheet per supplier + summary ───────────────────────────────

def create_format_a(wb, data):
    """Format A: one sheet per supplier + global summary sheet."""
    suppliers = data["suppliers"]
    categories = data.get("categories", [])
    requirements = data.get("requirements", [])
    responses_data = data.get("responses", {})
    project_name = data.get("project", "RFP Evaluation")

    # Build category lookup
    cat_by_code = {}
    for c in categories:
        cat_by_code[c["code"]] = c
        cat_by_code[c["title"]] = c
        cat_by_code[c["id"]] = c

    flat_cats, children_map = build_category_tree(categories)

    # Group requirements by category_name
    reqs_by_cat = defaultdict(list)
    for req in requirements:
        reqs_by_cat[req.get("category_name", "")].append(req)

    def req_category_ids(req):
        """Find category ID for a requirement."""
        name = req.get("category_name", "")
        cat = cat_by_code.get(name)
        return cat["id"] if cat else None

    # Build req → category_id mapping
    req_to_cat_id = {}
    for req in requirements:
        cid = req_category_ids(req)
        if cid:
            req_to_cat_id[req["code"]] = cid

    # ── Summary sheet (created first, filled last) ──
    ws_summary = wb.active
    ws_summary.title = "Synthèse"

    # ── Per-supplier sheets ──
    supplier_category_scores = {}  # {supplier: {cat_id: weighted_score}}
    supplier_global_scores = {}    # {supplier: global_weighted_score}

    for supplier in suppliers:
        ws = wb.create_sheet(title=supplier[:31])  # Excel sheet name max 31 chars
        supplier_responses = responses_data.get(supplier, [])
        resp_by_req = {r["requirement_id_external"]: r for r in supplier_responses}

        cat_scores = {}

        # Title
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"{project_name} — Fournisseur : {supplier}"
        title_cell.font = Font(bold=True, color=COLORS["white"], size=14)
        title_cell.fill = PatternFill("solid", fgColor=COLORS["header_dark"])
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Column headers (row 2)
        headers = ["Catégorie", "Code", "Exigence", "Poids", "Oblig.", "Score", "Statut", "Commentaire", "Question soutenance"]
        col_widths = [30, 10, 40, 8, 8, 8, 14, 60, 40]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True, color=COLORS["white"])
            cell.fill = PatternFill("solid", fgColor=COLORS["header_mid"])
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[2].height = 20
        ws.freeze_panes = "A3"

        row = 3
        for cat, depth in flat_cats:
            cat_reqs = [r for r in requirements if req_to_cat_id.get(r["code"]) == cat["id"]]
            if not cat_reqs:
                continue

            # Category header row
            ws.merge_cells(f"A{row}:I{row}")
            cat_cell = ws[f"A{row}"]
            indent = "  " * depth
            cat_cell.value = f"{indent}{'■ ' if depth == 0 else '▸ '}{cat['title']} ({len(cat_reqs)} exigences)"
            cat_cell.font = Font(bold=True, color=COLORS["white"], size=11 if depth == 0 else 10)
            cat_cell.fill = PatternFill("solid", fgColor=COLORS["header_mid"] if depth == 0 else "4A90D9")
            cat_cell.alignment = Alignment(vertical="center", indent=depth)
            ws.row_dimensions[row].height = 22
            row += 1

            # Requirements rows
            for i, req in enumerate(cat_reqs):
                resp = resp_by_req.get(req["code"], {})
                score = resp.get("ai_score")
                status = resp.get("status", "pending")
                comment = resp.get("ai_comment", "")
                question = resp.get("question", "")
                is_mandatory = req.get("is_mandatory", False)

                row_fill = PatternFill("solid", fgColor=COLORS["mandatory_bg"] if is_mandatory else (COLORS["row_alt"] if i % 2 else COLORS["white"]))

                values = [
                    "",  # Category column (empty, shown in header)
                    req["code"],
                    req["title"],
                    req.get("weight", ""),
                    "🔴 Oui" if is_mandatory else "",
                    score if score is not None else "—",
                    status_label(status, score, is_mandatory),
                    comment,
                    question,
                ]

                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if col_idx not in (6,):  # Score cell gets special treatment
                        cell.fill = row_fill
                    else:
                        cell.fill = score_fill(score)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Thin border
                    thin = Side(style="thin", color="CCCCCC")
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                ws.row_dimensions[row].height = 40 if len(comment) > 100 else 25
                row += 1

            # Category subtotal
            ws_score, _, evaluated = weighted_score(cat_reqs, resp_by_req)
            cat_scores[cat["id"]] = ws_score

            ws.merge_cells(f"A{row}:E{row}")
            subtotal_cell = ws[f"A{row}"]
            subtotal_cell.value = f"Score pondéré : {cat['title']} ({evaluated}/{len(cat_reqs)} exigences évaluées)"
            subtotal_cell.font = Font(bold=True, italic=True)
            subtotal_cell.fill = PatternFill("solid", fgColor=COLORS["summary_total"])

            score_cell = ws.cell(row=row, column=6, value=ws_score if ws_score is not None else "—")
            score_cell.font = Font(bold=True)
            score_cell.fill = score_fill(ws_score) if ws_score else PatternFill("solid", fgColor=COLORS["pending"])
            score_cell.alignment = Alignment(horizontal="center")

            row += 2  # Blank separator row

        # Global score
        global_score, _, total_evaluated = weighted_score(requirements, resp_by_req)
        supplier_global_scores[supplier] = global_score
        supplier_category_scores[supplier] = cat_scores

        ws.merge_cells(f"A{row}:E{row}")
        global_cell = ws[f"A{row}"]
        global_cell.value = f"SCORE GLOBAL PONDÉRÉ ({total_evaluated}/{len(requirements)} exigences évaluées)"
        global_cell.font = Font(bold=True, size=12, color=COLORS["white"])
        global_cell.fill = PatternFill("solid", fgColor=COLORS["header_dark"])

        global_score_cell = ws.cell(row=row, column=6, value=global_score if global_score is not None else "—")
        global_score_cell.font = Font(bold=True, size=12, color=COLORS["white"])
        global_score_cell.fill = PatternFill("solid", fgColor=COLORS["header_dark"])
        global_score_cell.alignment = Alignment(horizontal="center")

        ws.row_dimensions[row].height = 28

    # ── Fill summary sheet ──
    _fill_summary_sheet(ws_summary, project_name, suppliers, flat_cats, requirements,
                        req_to_cat_id, supplier_category_scores, supplier_global_scores,
                        responses_data)


def _fill_summary_sheet(ws, project_name, suppliers, flat_cats, requirements,
                        req_to_cat_id, supplier_category_scores, supplier_global_scores,
                        responses_data):
    """Fill the summary/synthesis sheet."""

    # Title
    n_cols = 3 + len(suppliers) * 2
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    title = ws["A1"]
    title.value = f"SYNTHÈSE ÉVALUATION — {project_name}"
    title.font = Font(bold=True, color=COLORS["white"], size=14)
    title.fill = PatternFill("solid", fgColor=COLORS["header_dark"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Global ranking header
    row = 3
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"].value = "CLASSEMENT GÉNÉRAL"
    ws[f"A{row}"].font = Font(bold=True, size=12, color=COLORS["white"])
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=COLORS["header_mid"])

    row += 1
    ws.cell(row=row, column=1, value="Rang").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Fournisseur").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Score global pondéré").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Alertes éliminatoires").font = Font(bold=True)

    row += 1
    ranked = sorted(
        [(s, supplier_global_scores.get(s)) for s in suppliers],
        key=lambda x: (x[1] is None, -(x[1] or 0))
    )

    for rank, (supplier, score) in enumerate(ranked, 1):
        # Count mandatory fails
        resp_list = responses_data.get(supplier, [])
        resp_by_req = {r["requirement_id_external"]: r for r in resp_list}
        mandatory_fails = [
            req["code"] for req in requirements
            if req.get("is_mandatory") and
            resp_by_req.get(req["code"], {}).get("ai_score", None) is not None and
            resp_by_req.get(req["code"], {}).get("ai_score") < 3
        ]

        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=supplier)
        score_cell = ws.cell(row=row, column=3, value=score if score is not None else "—")
        score_cell.fill = score_fill(score)
        score_cell.alignment = Alignment(horizontal="center")

        alert_cell = ws.cell(row=row, column=4, value=", ".join(mandatory_fails) if mandatory_fails else "—")
        if mandatory_fails:
            alert_cell.font = Font(color=COLORS["alert_red"], bold=True)

        row += 1

    row += 2

    # Scores by category
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"].value = "SCORES PAR CATÉGORIE"
    ws[f"A{row}"].font = Font(bold=True, size=12, color=COLORS["white"])
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=COLORS["header_mid"])
    ws.merge_cells(f"C{row}:{get_column_letter(2 + len(suppliers))}{row}")
    row += 1

    # Category score header
    ws.cell(row=row, column=1, value="Catégorie").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Nb exigences").font = Font(bold=True)
    for i, supplier in enumerate(suppliers):
        ws.cell(row=row, column=3 + i, value=supplier).font = Font(bold=True)
        ws.cell(row=row, column=3 + i).fill = PatternFill("solid", fgColor=COLORS["header_light"])
        ws.cell(row=row, column=3 + i).alignment = Alignment(horizontal="center")

    row += 1
    for cat, depth in flat_cats:
        cat_reqs = [r for r in requirements if req_to_cat_id.get(r["code"]) == cat["id"]]
        if not cat_reqs:
            continue

        indent = "  " * depth
        ws.cell(row=row, column=1, value=f"{indent}{cat['title']}")
        ws.cell(row=row, column=2, value=len(cat_reqs))

        for i, supplier in enumerate(suppliers):
            score = supplier_category_scores.get(supplier, {}).get(cat["id"])
            cell = ws.cell(row=row, column=3 + i, value=score if score is not None else "—")
            cell.fill = score_fill(score)
            cell.alignment = Alignment(horizontal="center")

        row += 1

    # Global row
    ws.cell(row=row, column=1, value="SCORE GLOBAL PONDÉRÉ").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(requirements))
    for i, supplier in enumerate(suppliers):
        score = supplier_global_scores.get(supplier)
        cell = ws.cell(row=row, column=3 + i, value=score if score is not None else "—")
        cell.fill = score_fill(score)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row += 3

    # Soutenance questions summary
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"].value = "QUESTIONS POUR LA SOUTENANCE"
    ws[f"A{row}"].font = Font(bold=True, size=12, color=COLORS["white"])
    ws[f"A{row}"].fill = PatternFill("solid", fgColor=COLORS["header_dark"])
    row += 1

    ws.cell(row=row, column=1, value="Fournisseur").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Exigence").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Score").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Question").font = Font(bold=True)
    row += 1

    for supplier in suppliers:
        resp_list = responses_data.get(supplier, [])
        questions = [(r, req) for r in resp_list
                     if r.get("question")
                     for req in requirements
                     if req["code"] == r["requirement_id_external"]]

        # Sort by score ascending (most critical first)
        questions.sort(key=lambda x: (x[0].get("ai_score") or 99))

        for resp, req in questions:
            ws.cell(row=row, column=1, value=supplier)
            ws.cell(row=row, column=2, value=f"{req['code']} — {req['title']}")
            score_cell = ws.cell(row=row, column=3, value=resp.get("ai_score"))
            score_cell.fill = score_fill(resp.get("ai_score"))
            score_cell.alignment = Alignment(horizontal="center")
            question_cell = ws.cell(row=row, column=4, value=resp.get("question", ""))
            question_cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 35
            row += 1

    # Column widths for summary
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    for i in range(len(suppliers)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 20
    if ws.max_column >= 4:
        ws.column_dimensions["D"].width = 60


# ─── Format B: Single consolidated sheet ─────────────────────────────────────

def create_format_b(wb, data):
    """Format B: single sheet with all suppliers as columns."""
    suppliers = data["suppliers"]
    categories = data.get("categories", [])
    requirements = data.get("requirements", [])
    responses_data = data.get("responses", {})
    project_name = data.get("project", "RFP Evaluation")

    flat_cats, _ = build_category_tree(categories)

    # Build lookup
    cat_by_code = {}
    for c in categories:
        cat_by_code[c["code"]] = c
        cat_by_code[c["title"]] = c
        cat_by_code[c["id"]] = c

    req_to_cat_id = {}
    for req in requirements:
        name = req.get("category_name", "")
        cat = cat_by_code.get(name)
        if cat:
            req_to_cat_id[req["code"]] = cat["id"]

    ws = wb.active
    ws.title = "Grille comparative"

    # Title row
    n_supplier_cols = len(suppliers) * 2  # score + comment per supplier
    total_cols = 4 + n_supplier_cols  # base cols + supplier cols
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    title = ws["A1"]
    title.value = f"GRILLE D'ÉVALUATION COMPARATIVE — {project_name}"
    title.font = Font(bold=True, color=COLORS["white"], size=13)
    title.fill = PatternFill("solid", fgColor=COLORS["header_dark"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Supplier header (row 2) — merged cells per supplier
    for i, supplier in enumerate(suppliers):
        col_start = 5 + i * 2
        ws.merge_cells(f"{get_column_letter(col_start)}2:{get_column_letter(col_start + 1)}2")
        cell = ws.cell(row=2, column=col_start, value=supplier)
        cell.font = Font(bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_mid"])
        cell.alignment = Alignment(horizontal="center")

    # Column headers (row 3)
    base_headers = ["Catégorie", "Code", "Exigence", "Poids", "Oblig."]
    for col_idx, header in enumerate(base_headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=COLORS["header_mid"])
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, supplier in enumerate(suppliers):
        col_score = 5 + i * 2
        col_comment = col_score + 1
        cell_s = ws.cell(row=3, column=col_score, value="Score")
        cell_s.font = Font(bold=True)
        cell_s.fill = PatternFill("solid", fgColor=COLORS["header_light"])
        cell_s.alignment = Alignment(horizontal="center")

        cell_c = ws.cell(row=3, column=col_comment, value="Commentaire / Question soutenance")
        cell_c.font = Font(bold=True)
        cell_c.fill = PatternFill("solid", fgColor=COLORS["header_light"])
        cell_c.alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A4"
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    for i in range(len(suppliers)):
        ws.column_dimensions[get_column_letter(6 + i * 2 - 1)].width = 8
        ws.column_dimensions[get_column_letter(6 + i * 2)].width = 45

    # Build response lookup per supplier
    all_resp = {}
    for supplier in suppliers:
        resp_list = responses_data.get(supplier, [])
        all_resp[supplier] = {r["requirement_id_external"]: r for r in resp_list}

    row = 4
    for cat, depth in flat_cats:
        cat_reqs = [r for r in requirements if req_to_cat_id.get(r["code"]) == cat["id"]]
        if not cat_reqs:
            continue

        # Category header
        ws.merge_cells(f"A{row}:{get_column_letter(total_cols)}{row}")
        cat_cell = ws[f"A{row}"]
        indent = "  " * depth
        cat_cell.value = f"{indent}{'■ ' if depth == 0 else '▸ '}{cat['title']} ({len(cat_reqs)} exigences)"
        cat_cell.font = Font(bold=True, color=COLORS["white"], size=11 if depth == 0 else 10)
        cat_cell.fill = PatternFill("solid", fgColor=COLORS["header_mid"] if depth == 0 else "4A90D9")
        ws.row_dimensions[row].height = 22
        row += 1

        for i_req, req in enumerate(cat_reqs):
            is_mandatory = req.get("is_mandatory", False)
            base_fill = PatternFill("solid", fgColor=COLORS["mandatory_bg"] if is_mandatory else (COLORS["row_alt"] if i_req % 2 else COLORS["white"]))

            values = [
                "",
                req["code"],
                req["title"],
                req.get("weight", ""),
                "🔴" if is_mandatory else "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.fill = base_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                thin = Side(style="thin", color="CCCCCC")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for i, supplier in enumerate(suppliers):
                resp = all_resp[supplier].get(req["code"], {})
                score = resp.get("ai_score")
                status = resp.get("status", "pending")
                comment = resp.get("ai_comment", "")
                question = resp.get("question", "")

                col_score = 5 + i * 2 + 1
                col_comment = col_score + 1

                score_cell = ws.cell(row=row, column=col_score, value=score if score is not None else "—")
                score_cell.fill = score_fill(score)
                score_cell.font = Font(bold=True)
                score_cell.alignment = Alignment(horizontal="center", vertical="center")
                thin = Side(style="thin", color="CCCCCC")
                score_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                comment_text = comment
                if question:
                    comment_text += f"\n\n❓ {question}"
                comment_cell = ws.cell(row=row, column=col_comment, value=comment_text)
                comment_cell.alignment = Alignment(wrap_text=True, vertical="top")
                comment_cell.fill = base_fill
                comment_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            max_comment_len = max(
                len(all_resp[s].get(req["code"], {}).get("ai_comment", ""))
                for s in suppliers
            )
            ws.row_dimensions[row].height = min(max(25, max_comment_len // 4), 120)
            row += 1

        row += 1  # blank separator

    # Global scores row
    ws.cell(row=row, column=1, value="SCORE GLOBAL PONDÉRÉ").font = Font(bold=True)
    for i, supplier in enumerate(suppliers):
        resp_by_req = all_resp[supplier]
        global_score, _, _ = weighted_score(requirements, resp_by_req)
        col_score = 5 + i * 2 + 1
        cell = ws.cell(row=row, column=col_score, value=global_score if global_score is not None else "—")
        cell.fill = score_fill(global_score)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 28


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate RFP evaluation Excel grid")
    parser.add_argument("data_json", help="Input JSON file with evaluation data")
    parser.add_argument("output_xlsx", help="Output Excel file path")
    parser.add_argument(
        "--format", choices=["A", "B"], default="A",
        help="A=one sheet per supplier + summary (default), B=single consolidated sheet"
    )
    args = parser.parse_args()

    # Load data
    data_path = Path(args.data_json)
    if not data_path.exists():
        print(f"ERROR: Input file not found: {args.data_json}")
        sys.exit(1)

    with open(data_path, encoding="utf-8") as f:
        data = json.load(f)

    # Validate minimal structure
    required_keys = ["suppliers", "requirements", "responses"]
    for key in required_keys:
        if key not in data:
            print(f"ERROR: Missing required key in JSON: '{key}'")
            sys.exit(1)

    if not data["suppliers"]:
        print("ERROR: No suppliers defined in data.")
        sys.exit(1)

    # Create workbook
    wb = openpyxl.Workbook()

    if args.format == "A":
        print(f"Generating Format A (per-supplier sheets + summary)...")
        create_format_a(wb, data)
    else:
        print(f"Generating Format B (consolidated single sheet)...")
        create_format_b(wb, data)

    # Save
    output_path = Path(args.output_xlsx)
    wb.save(output_path)
    print(f"✅ Excel file generated: {output_path}")
    print(f"   Suppliers: {', '.join(data['suppliers'])}")
    print(f"   Requirements: {len(data.get('requirements', []))}")
    print(f"   Categories: {len(data.get('categories', []))}")


if __name__ == "__main__":
    main()
